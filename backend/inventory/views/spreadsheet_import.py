"""
Spreadsheet import views for bulk adding new inventory items.

This module handles Excel import functionality for bulk item creation,
including template generation and import processing using the new tagging system.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db import transaction
from decimal import Decimal, InvalidOperation

from inventory.models import Item, StockItem, Organization, Tag, TagGroup
from .utils import audit_log_state, audit_log_event


def find_or_create_tag(tag_name, general_tag_group):
    """
    Find an existing tag by name (case-insensitive) or create a new one in the general group.
    
    Args:
        tag_name: Name of the tag to find or create
        general_tag_group: TagGroup to create new tags in
        
    Returns:
        Tag: Found or created tag
    """
    # First try to find existing tag (case-insensitive)
    existing_tag = Tag.objects.filter(name__iexact=tag_name, is_active=True).first()
    if existing_tag:
        return existing_tag
    
    # Create new tag in the general group
    new_tag = Tag.objects.create(
        name=tag_name.strip(),
        tag_group=general_tag_group,
        is_active=True
    )
    return new_tag


@login_required
@permission_required('inventory.add_viaspreadsheet_item', raise_exception=True)
def download_import_template(request):
    """
    Generate and download an Excel template for importing new items.
    
    Returns:
        HttpResponse: Excel file download with template headers and sample data
    """
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Import Template"
    
    # Headers for import
    headers = [
        'Item Name', 'Manufacturer', 'GTIN', 'Tags', 
        'Items Per Box', 'Cost Per Item', 'URL', 'Public Notes', 'Private Notes',
        'Organization', 'Quantity', 'Location', 'Detail', 'Date Received', 
        'Expiration Date', 'Lot Number', 'Stock Notes'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample data row for guidance
    sample_data = [
        'Medical Gloves', 'Nitrile Brand', '1234567890123', 'PPE, Disposable', 
        '100', '0.15', 'https://example.com', 'Disposable nitrile gloves', 'Store in cool dry place',
        'UCSD Health', '1000', 'Storage Room A', 'Size Large', str(date.today()), 
        '', 'LOT2024001', 'Received in good condition'
    ]
    
    for col, data in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=data)
        cell.font = Font(italic=True, color='808080')
    
    # Set column widths for better readability
    column_widths = [15, 15, 15, 20, 12, 12, 20, 20, 20, 15, 10, 15, 15, 12, 12, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        "Item Import Template Instructions",
        "",
        "Required Fields:",
        "- Item Name: Unique name for the item",
        "- Tags: Comma-separated list of tag names (e.g., 'PPE, Disposable, Medical')",
        "- Organization: Must match existing organization name", 
        "- Quantity: Number of items (must be positive integer)",
        "- Location: Storage location for the stock",
        "- Date Received: Date in YYYY-MM-DD format",
        "",
        "Optional Fields:",
        "- Manufacturer: Brand or manufacturer name",
        "- GTIN: Global Trade Item Number (barcode)",
        "- Items Per Box: Number of items per package",
        "- Cost Per Item: Cost in dollars (e.g., 1.50)",
        "- URL: Product website or documentation link",
        "- Public Notes: Notes visible to all users",
        "- Private Notes: Notes visible only to MAI members",
        "- Detail: Variant details (size, color, etc.)",
        "- Expiration Date: For perishable items (YYYY-MM-DD)",
        "- Lot Number: Batch identifier",
        "- Stock Notes: Notes specific to this stock entry",
        "",
        "Tags Information:",
        "- Use comma-separated tag names: 'PPE, Disposable, Medical'",
        "- Tag names are case-insensitive and will be automatically matched",
        "- If a tag doesn't exist, it will be created in the 'General' tag group",
        "- Multiple items can share the same tags",
        "",
        "Important Notes:",
        "- If an item name already exists, new stock will be added to the existing item",
        "- When reusing existing items, only stock details (quantity, location, etc.) are used",
        "- GTIN placement: If Detail field is empty, GTIN goes on the Item; if Detail has content, GTIN goes on the Stock Item",
        "- GTINs must be unique within their respective location (Item or Stock Item)",
        "- Organizations must already exist in the system",
        "- Dates should be in YYYY-MM-DD format",
        "- Delete the sample row before importing",
        "",
        "Available Tags (by Tag Group):",
    ]
    
    # Add tag group and tag information
    tag_groups = TagGroup.objects.prefetch_related('tags').filter(is_active=True).order_by('sort_order', 'name')
    for tag_group in tag_groups:
        instructions.append(f"  {tag_group.name}:")
        for tag in tag_group.tags.filter(is_active=True).order_by('sort_order', 'name'):
            instructions.append(f"    - {tag.name}")
    
    instructions.append("")
    instructions.append("Available Organizations:")
    organizations = Organization.objects.all().order_by('name')
    for org in organizations:
        instructions.append(f"  - {org.name}")
    
    # Add instructions to sheet
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction)
        if row == 1:  # Title
            cell.font = Font(bold=True, size=14)
        elif instruction.endswith(":") and not instruction.startswith(" "):  # Headers
            cell.font = Font(bold=True)
    
    # Set column width for instructions
    instructions_ws.column_dimensions['A'].width = 80
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'item_import_template_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


@login_required
@permission_required('inventory.add_viaspreadsheet_item', raise_exception=True)
def upload_spreadsheet(request):
    """
    Handle Excel file upload to create new items and stock items.
    
    Returns:
        HttpResponse: Rendered template or redirect
    """
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file was uploaded.')
            # Prepare context for the template
            tag_groups = TagGroup.objects.prefetch_related('tags').filter(is_active=True).order_by('sort_order', 'name')
            organizations = Organization.objects.all().order_by('name')
            
            context = {
                'tag_groups': tag_groups,
                'organizations': organizations,
            }
            return render(request, 'spreadsheet_import/upload.html', context)
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
            # Prepare context for the template
            tag_groups = TagGroup.objects.prefetch_related('tags').filter(is_active=True).order_by('sort_order', 'name')
            organizations = Organization.objects.all().order_by('name')
            
            context = {
                'tag_groups': tag_groups,
                'organizations': organizations,
            }
            return render(request, 'spreadsheet_import/upload.html', context)
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Validate headers
            expected_headers = [
                'Item Name', 'Manufacturer', 'GTIN', 'Tags', 
                'Items Per Box', 'Cost Per Item', 'URL', 'Public Notes', 'Private Notes',
                'Organization', 'Quantity', 'Location', 'Detail', 'Date Received', 
                'Expiration Date', 'Lot Number', 'Stock Notes'
            ]
            
            if ws.max_row < 1:
                messages.error(request, 'The uploaded file appears to be empty.')
                # Prepare context for the template
                tag_groups = TagGroup.objects.prefetch_related('tags').all().order_by('sort_order', 'name')
                organizations = Organization.objects.all().order_by('name')
                
                context = {
                    'tag_groups': tag_groups,
                    'organizations': organizations,
                }
                return render(request, 'spreadsheet_import/upload.html', context)
                
            actual_headers = [cell.value for cell in ws[1]]
            if actual_headers != expected_headers:
                messages.error(request, 'Invalid file format. Please use the template from the download function.')
                # Prepare context for the template
                tag_groups = TagGroup.objects.prefetch_related('tags').all().order_by('sort_order', 'name')
                organizations = Organization.objects.all().order_by('name')
                
                context = {
                    'tag_groups': tag_groups,
                    'organizations': organizations,
                }
                return render(request, 'spreadsheet_import/upload.html', context)
            
            # Process data rows
            error_count = 0
            errors = []
            
            # Prefetch all data to avoid N+1 queries in the loop
            organizations_dict = {org.name.lower(): org for org in Organization.objects.all()}
            # Create tag lookup dictionary: {tag_name_lower: tag_obj}
            tags_dict = {tag.name.lower(): tag for tag in Tag.objects.all()}
            
            # First pass: validate all data
            items_to_create = []
            # Track items being created in this import to handle duplicates within the same file
            items_in_current_import = {}  # name_lower -> item_data
            
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Extract row data
                    item_name = str(ws.cell(row=row_num, column=1).value or '').strip()
                    manufacturer = str(ws.cell(row=row_num, column=2).value or '').strip()
                    gtin = str(ws.cell(row=row_num, column=3).value or '').strip()
                    tags_str = str(ws.cell(row=row_num, column=4).value or '').strip()
                    items_per_box = ws.cell(row=row_num, column=5).value
                    cost_per_item = ws.cell(row=row_num, column=6).value
                    url = str(ws.cell(row=row_num, column=7).value or '').strip()
                    public_notes = str(ws.cell(row=row_num, column=8).value or '').strip()
                    private_notes = str(ws.cell(row=row_num, column=9).value or '').strip()
                    organization_name = str(ws.cell(row=row_num, column=10).value or '').strip()
                    quantity = ws.cell(row=row_num, column=11).value
                    location = str(ws.cell(row=row_num, column=12).value or '').strip()
                    detail = str(ws.cell(row=row_num, column=13).value or '').strip()
                    date_received = ws.cell(row=row_num, column=14).value
                    expiration_date = ws.cell(row=row_num, column=15).value
                    lot_number = str(ws.cell(row=row_num, column=16).value or '').strip()
                    stock_notes = str(ws.cell(row=row_num, column=17).value or '').strip()
                    
                    # Skip empty rows
                    if not item_name:
                        continue
                    
                    # Parse tags (comma-separated)
                    tag_names = []
                    if tags_str:
                        tag_names = [name.strip() for name in tags_str.split(',') if name.strip()]
                    
                    # Validate required fields
                    if not organization_name:
                        errors.append(f"Row {row_num}: Organization is required")
                        error_count += 1
                        continue
                        
                    if not quantity:
                        errors.append(f"Row {row_num}: Quantity is required")
                        error_count += 1
                        continue
                        
                    if not location:
                        errors.append(f"Row {row_num}: Location is required")
                        error_count += 1
                        continue
                        
                    if not date_received:
                        errors.append(f"Row {row_num}: Date Received is required")
                        error_count += 1
                        continue
                    
                    item_name_lower = item_name.lower()
                    
                    # Check if item already exists in database
                    existing_item = Item.objects.filter(name__iexact=item_name).first()
                    
                    # Check if item is being created in this import
                    item_from_current_import = items_in_current_import.get(item_name_lower)
                    
                    # Determine which item to use (existing in DB, or from current import)
                    reference_item_data = None
                    if existing_item:
                        reference_item_data = {
                            'tag_names': [tag.name for tag in existing_item.tags.all()]
                        }
                    elif item_from_current_import:
                        reference_item_data = {
                            'tag_names': item_from_current_import['tag_names']
                        }
                    
                    # If we have a reference (existing or from current import), validate consistency
                    if reference_item_data:
                        # Check for tag consistency - warn if tags differ but don't fail
                        existing_tag_names = set(tag.lower() for tag in reference_item_data['tag_names'])
                        new_tag_names = set(tag.lower() for tag in tag_names)
                        if existing_tag_names != new_tag_names:
                            existing_source = "database" if existing_item else "earlier in this import"
                            existing_tags_str = ', '.join(reference_item_data['tag_names'])
                            new_tags_str = ', '.join(tag_names)
                            messages.warning(request, f"Row {row_num}: Item '{item_name}' exists in {existing_source} with tags '{existing_tags_str}', but row specifies '{new_tags_str}'. Using existing tags.")
                    
                    # Validate GTIN uniqueness if provided
                    if gtin:
                        if len(gtin) > 14:
                            errors.append(f"Row {row_num}: GTIN must be at most 14 characters")
                            error_count += 1
                            continue
                        
                        # Determine where GTIN will be placed based on detail field
                        gtin_on_stock_item = bool(detail.strip())
                        
                        if gtin_on_stock_item:
                            # GTIN goes on StockItem - check for conflicts in StockItem table
                            if StockItem.objects.filter(gtin=gtin).exists():
                                errors.append(f"Row {row_num}: GTIN '{gtin}' already exists on a stock item")
                                error_count += 1
                                continue
                        else:
                            # GTIN goes on Item - check for conflicts in Item table (excluding existing item if reusing)
                            gtin_conflict = Item.objects.filter(gtin=gtin).exclude(id=existing_item.id if existing_item else None).exists()
                            if gtin_conflict:
                                errors.append(f"Row {row_num}: GTIN '{gtin}' already exists on an item")
                                error_count += 1
                                continue
                    
                    # Validate tags using pre-fetched dictionary
                    valid_tags = []
                    invalid_tags = False
                    for tag_name in tag_names:
                        tag = tags_dict.get(tag_name.lower())
                        if not tag:
                            errors.append(f"Row {row_num}: Tag '{tag_name}' not found")
                            error_count += 1
                            invalid_tags = True
                        else:
                            valid_tags.append(tag)
                    
                    # Skip this row if any tags were invalid
                    if invalid_tags:
                        continue
                    
                    # Validate organization using pre-fetched dictionary
                    organization = organizations_dict.get(organization_name.lower())
                    if not organization:
                        errors.append(f"Row {row_num}: Organization '{organization_name}' not found")
                        error_count += 1
                        continue
                    
                    # Validate quantity
                    try:
                        quantity = int(quantity)
                        if quantity <= 0:
                            errors.append(f"Row {row_num}: Quantity must be positive")
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Quantity must be a valid number")
                        error_count += 1
                        continue
                    
                    # Validate items_per_box if provided
                    if items_per_box:
                        try:
                            items_per_box = int(items_per_box)
                            if items_per_box <= 0:
                                items_per_box = None
                        except (ValueError, TypeError):
                            items_per_box = None
                    else:
                        items_per_box = None
                    
                    # Validate cost_per_item if provided
                    if cost_per_item:
                        try:
                            cost_per_item = Decimal(str(cost_per_item))
                            if cost_per_item < 0:
                                cost_per_item = None
                        except (InvalidOperation, ValueError, TypeError):
                            cost_per_item = None
                    else:
                        cost_per_item = None
                    
                    # Validate date_received
                    if isinstance(date_received, str):
                        try:
                            date_received = datetime.strptime(date_received, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Date Received must be in YYYY-MM-DD format")
                            error_count += 1
                            continue
                    elif not isinstance(date_received, date):
                        try:
                            date_received = date_received.date() if hasattr(date_received, 'date') else date.today()
                        except Exception:
                            errors.append(f"Row {row_num}: Invalid Date Received format")
                            error_count += 1
                            continue
                    
                    # Validate expiration_date if provided
                    if expiration_date:
                        if isinstance(expiration_date, str):
                            try:
                                expiration_date = datetime.strptime(expiration_date, '%Y-%m-%d').date()
                            except ValueError:
                                errors.append(f"Row {row_num}: Expiration Date must be in YYYY-MM-DD format")
                                error_count += 1
                                continue
                        elif not isinstance(expiration_date, date):
                            try:
                                expiration_date = expiration_date.date() if hasattr(expiration_date, 'date') else None
                            except Exception:
                                expiration_date = None
                    else:
                        expiration_date = None
                    
                    # Process tags
                    tag_objects = []
                    if tags_str:
                        tag_names = [tag.strip() for tag in tags_str.split(',') if tag.strip()]
                        for tag_name in tag_names:
                            tag_obj = find_or_create_tag(tag_name, general_tag_group)
                            tag_objects.append(tag_obj)
                    
                    # Determine where GTIN should be placed based on detail field
                    gtin_on_stock_item = bool(detail.strip())
                    item_gtin = '' if gtin_on_stock_item else gtin
                    stock_gtin = gtin if gtin_on_stock_item else ''
                    
                    # Determine if we need to create a new item or reuse existing
                    will_create_new_item = not existing_item and not item_from_current_import
                    
                    # If creating new item, track it for subsequent rows
                    if will_create_new_item:
                        items_in_current_import[item_name_lower] = {
                            'tag_names': tag_names,
                            'item_data': {
                                'name': item_name,
                                'manufacturer': manufacturer,
                                'gtin': item_gtin,
                                'items_per_box': items_per_box,
                                'cost_per_item': cost_per_item,
                                'url': url,
                                'notes_public': public_notes,
                                'notes_private': private_notes,
                            },
                            'tags': tag_objects,
                        }
                    
                    # Store validated data for creation
                    items_to_create.append({
                        'row_num': row_num,
                        'existing_item': existing_item,  # None if creating new item
                        'item_name_lower': item_name_lower,  # For tracking within import
                        'valid_tags': valid_tags,  # Tags to assign to the item
                        'item_data': {
                            'name': item_name,
                            'manufacturer': manufacturer,
                            'gtin': item_gtin,
                            'items_per_box': items_per_box,
                            'cost_per_item': cost_per_item,
                            'url': url,
                            'notes_public': public_notes,
                            'notes_private': private_notes,
                        },
                        'tags': tag_objects,
                        'stock_data': {
                            'organization': organization,
                            'quantity': quantity,
                            'location': location,
                            'detail': detail,
                            'date_received': date_received,
                            'expiration_date': expiration_date,
                            'lot_number': lot_number,
                            'notes': stock_notes,
                            'gtin': stock_gtin,
                        }
                    })
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: Error processing row - {str(e)}")
                    error_count += 1
            
            # If there are validation errors, don't proceed with creation
            if error_count > 0:
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... and {len(errors) - 10} more errors.')
                messages.error(request, f'Found {error_count} validation errors. No items were created.')
                return redirect('spreadsheet_import_upload')
            
            # Create all items within a transaction
            created_item_count = 0
            added_stock_count = 0
            # Track items created in this transaction
            created_items_map = {}  # name_lower -> Item instance
            
            try:
                with transaction.atomic():
                    for item_data in items_to_create:
                        item_name_lower = item_data['item_name_lower']
                        
                        # Determine which item to use
                        if item_data['existing_item']:
                            # Use existing item from database
                            item = item_data['existing_item']
                        elif item_name_lower in created_items_map:
                            # Use item created earlier in this transaction
                            item = created_items_map[item_name_lower]
                        else:
                            # Create new Item
                            item = Item(**item_data['item_data'])
                            item.save()
                            
                            # Assign tags to the new item
                            if item_data['valid_tags']:
                                item.tags.set(item_data['valid_tags'])
                            
                            created_item_count += 1
                            created_items_map[item_name_lower] = item
                            
                            # Log item creation
                            audit_log_event(
                                request.user,
                                f"Created item \"{item.name}\" via spreadsheet import",
                                audit_log_state(None),
                                audit_log_state(item)
                            )
                        
                        # Create StockItem (always create new stock)
                        stock_item = StockItem(
                            item=item,
                            **item_data['stock_data']
                        )
                        stock_item.save()
                        added_stock_count += 1
                        
                        # Log stock item creation
                        action_desc = "Added initial stock" if item_name_lower in created_items_map else "Added stock"
                        audit_log_event(
                            request.user,
                            f"{action_desc} for \"{item.name}\" via spreadsheet import - {stock_item.quantity} units to {stock_item.location}",
                            audit_log_state(None),
                            audit_log_state(stock_item)
                        )
                
                # Build success message
                if created_item_count > 0 and added_stock_count > created_item_count:
                    messages.success(request, f'Successfully created {created_item_count} new items and added stock to {added_stock_count - created_item_count} existing items from spreadsheet.')
                elif created_item_count > 0:
                    messages.success(request, f'Successfully created {created_item_count} new items from spreadsheet.')
                else:
                    messages.success(request, f'Successfully added stock to {added_stock_count} existing items from spreadsheet.')
                    
                return redirect('spreadsheet_import_upload')
                
            except Exception as e:
                messages.error(request, f'Error creating items: {str(e)}')
                return redirect('spreadsheet_import_upload')
        
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            # Prepare context for the template
            tag_groups = TagGroup.objects.prefetch_related('tags').all().order_by('sort_order', 'name')
            organizations = Organization.objects.all().order_by('name')
            
            context = {
                'tag_groups': tag_groups,
                'organizations': organizations,
            }
            return render(request, 'spreadsheet_import/upload.html', context)
    
    # Prepare context for the template
    tag_groups = TagGroup.objects.prefetch_related('tags').all().order_by('sort_order', 'name')
    organizations = Organization.objects.all().order_by('name')
    
    context = {
        'tag_groups': tag_groups,
        'organizations': organizations,
    }
    
    return render(request, 'spreadsheet_import/upload.html', context)