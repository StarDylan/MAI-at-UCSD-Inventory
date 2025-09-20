"""
Bulk Edit views for the inventory application.

This module handles bulk editing functionality for inventory items,
including spreadsheet export, import, validation, and change execution
with proper audit logging.
"""

import json
import openpyxl
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from django.db import transaction

from inventory.models import Item, StockItem, Organization, Tag, TagGroup
from .utils import audit_log_state, audit_log_event


# Helper function to serialize datetime objects for JSON
def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")


# Helper function to parse ISO format dates from JSON
def parse_date_fields(data):
    """Parse ISO format dates in the data dictionary"""
    date_fields = ['date_received', 'expiration_date']
    
    for row in data:
        for field in date_fields:
            if field in row and row[field] and isinstance(row[field], str):
                try:
                    # Try to parse as date
                    row[field] = datetime.fromisoformat(row[field]).date()
                except ValueError:
                    # If parsing fails, leave as string
                    pass
    
    return data


class BulkEditView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Main view for bulk editing inventory items.
    
    Provides a page with options to download the current inventory as a spreadsheet,
    upload a modified spreadsheet, preview changes, and execute them.
    """
    template_name = 'bulk_edit/index.html'
    permission_required = 'inventory.bulkedit_item'
    
    def get_context_data(self, **kwargs):
        """Add extra context data to the template."""
        context = super().get_context_data(**kwargs)
        return context


@login_required
@permission_required('inventory.bulkedit_item', raise_exception=True)
def download_inventory_spreadsheet(request):
    """
    Generate and download an Excel spreadsheet with all inventory items and stock items.
    
    Returns:
        HttpResponse: Excel file download with all inventory data
    """
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Export"
    
    # Headers for export
    headers = [
        'Item ID', 'Item Name', 'Manufacturer', 'Item GTIN', 'Tags', 
        'Items Per Box', 'Cost Per Item', 'URL', 'Public Notes', 'Private Notes',
        'Stock Item ID', 'Organization', 'Quantity', 'Location', 'Detail', 
        'Stock GTIN', 'Date Received', 'Expiration Date', 'Lot Number', 'Stock Notes',
        'Surplus Status'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Fetch all active items with related stock items
    items = Item.active_objects.all().prefetch_related('stock_items', 'tags', 'stock_items__organization').order_by('name')
    
    # Add data rows
    row_num = 2
    for item in items:
        # Get comma-separated list of tag names
        tag_names = ', '.join(tag.name for tag in item.tags.all())
        
        # If item has no stock items, add a row with just item data
        if not item.stock_items.exists():
            ws.cell(row=row_num, column=1, value=str(item.id))
            ws.cell(row=row_num, column=2, value=item.name)
            ws.cell(row=row_num, column=3, value=item.manufacturer)
            ws.cell(row=row_num, column=4, value=item.gtin)
            ws.cell(row=row_num, column=5, value=tag_names)
            ws.cell(row=row_num, column=6, value=item.items_per_box)
            ws.cell(row=row_num, column=7, value=float(item.cost_per_item) if item.cost_per_item else None)
            ws.cell(row=row_num, column=8, value=item.url)
            ws.cell(row=row_num, column=9, value=item.notes_public)
            ws.cell(row=row_num, column=10, value=item.notes_private)
            row_num += 1
        else:
            # Add a row for each stock item
            for stock in item.stock_items.all():
                ws.cell(row=row_num, column=1, value=str(item.id))
                ws.cell(row=row_num, column=2, value=item.name)
                ws.cell(row=row_num, column=3, value=item.manufacturer)
                ws.cell(row=row_num, column=4, value=item.gtin)
                ws.cell(row=row_num, column=5, value=tag_names)
                ws.cell(row=row_num, column=6, value=item.items_per_box)
                ws.cell(row=row_num, column=7, value=float(item.cost_per_item) if item.cost_per_item else None)
                ws.cell(row=row_num, column=8, value=item.url)
                ws.cell(row=row_num, column=9, value=item.notes_public)
                ws.cell(row=row_num, column=10, value=item.notes_private)
                
                # Stock item data
                ws.cell(row=row_num, column=11, value=str(stock.id))
                ws.cell(row=row_num, column=12, value=stock.organization.name)
                ws.cell(row=row_num, column=13, value=stock.quantity)
                ws.cell(row=row_num, column=14, value=stock.location)
                ws.cell(row=row_num, column=15, value=stock.detail)
                ws.cell(row=row_num, column=16, value=stock.gtin)
                ws.cell(row=row_num, column=17, value=stock.date_received)
                ws.cell(row=row_num, column=18, value=stock.expiration_date)
                ws.cell(row=row_num, column=19, value=stock.lot_number)
                ws.cell(row=row_num, column=20, value=stock.notes)
                ws.cell(row=row_num, column=21, value=stock.surplus_status)
                
                row_num += 1
    
    # Set column widths for better readability
    column_widths = [36, 25, 15, 15, 25, 12, 12, 20, 20, 20, 36, 15, 10, 15, 15, 15, 12, 12, 12, 20, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        "Bulk Edit Instructions",
        "",
        "This spreadsheet contains all inventory items and their stock items. You can make the following changes:",
        "",
        "1. Edit item details (name, manufacturer, GTIN, tags, etc.)",
        "2. Edit stock item details (quantity, location, detail, etc.)",
        "3. Move stock items between items by changing the Item ID",
        "4. Combine items by moving all stock items to a single item",
        "",
        "Important Notes:",
        "- DO NOT change the Item ID or Stock Item ID columns unless you want to move stock items",
        "- To move a stock item to another item, change its Item ID to match the target item",
        "- To delete an item, remove all its stock items (move them or set quantity to 0)",
        "- Items with no stock items will be deleted if they have no stock items in the uploaded sheet",
        "- Tags must be comma-separated and must already exist in the system",
        "- Organizations must already exist in the system",
        "- Dates should be in YYYY-MM-DD format",
        "",
        "Available Tags (by Tag Group):",
    ]
    
    # Add tag group and tag information
    tag_groups = TagGroup.objects.prefetch_related('tags').filter(is_active=True).order_by('sort_order', 'name')
    for tag_group in tag_groups:
        instructions.append(f"  {tag_group.name}:")
        for tag in tag_group.tags.filter(is_active=True).order_by('sort_order', 'name'):
            instructions.append(f"    - {tag.name}")
    
    instructions.append("")
    instructions.append("Available Organizations:")
    organizations = Organization.objects.all().order_by('name')
    for org in organizations:
        instructions.append(f"  - {org.name}")
    
    # Add instructions to sheet
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction)
        if row == 1:  # Title
            cell.font = Font(bold=True, size=14)
        elif instruction.endswith(":") and not instruction.startswith(" "):  # Headers
            cell.font = Font(bold=True)
    
    # Set column width for instructions
    instructions_ws.column_dimensions['A'].width = 80
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventory_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


@login_required
@permission_required('inventory.bulkedit_item', raise_exception=True)
def upload_inventory_spreadsheet(request):
    """
    Handle Excel file upload to preview changes to inventory.
    
    Returns:
        HttpResponse: Rendered template with preview of changes
    """
    if request.method != 'POST':
        return redirect('bulk_edit')
        
    if 'excel_file' not in request.FILES:
        messages.error(request, 'No file was uploaded.')
        return redirect('bulk_edit')
    
    excel_file = request.FILES['excel_file']
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('bulk_edit')
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Validate headers
        expected_headers = [
            'Item ID', 'Item Name', 'Manufacturer', 'Item GTIN', 'Tags', 
            'Items Per Box', 'Cost Per Item', 'URL', 'Public Notes', 'Private Notes',
            'Stock Item ID', 'Organization', 'Quantity', 'Location', 'Detail', 
            'Stock GTIN', 'Date Received', 'Expiration Date', 'Lot Number', 'Stock Notes',
            'Surplus Status'
        ]
        
        if ws.max_row < 2:  # Header row + at least one data row
            messages.error(request, 'The uploaded file appears to be empty.')
            return redirect('bulk_edit')
            
        actual_headers = [cell.value for cell in ws[1]]
        if actual_headers != expected_headers:
            messages.error(request, 'Invalid file format. Please use the template from the download function.')
            return redirect('bulk_edit')
        
        # Process data rows to identify changes
        errors = []
        
        # Prefetch all data to avoid N+1 queries in the loop
        organizations_dict = {org.name.lower(): org for org in Organization.objects.all()}
        tags_dict = {tag.name.lower(): tag for tag in Tag.objects.all()}
        
        # Get all existing items and stock items
        existing_items = {str(item.id): item for item in Item.objects.all()}
        existing_stock_items = {str(stock.id): stock for stock in StockItem.objects.all()}
        
        # Store uploaded data for session
        uploaded_data = []
        
        for row_num in range(2, ws.max_row + 1):
            try:
                # Extract row data
                item_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                item_name = str(ws.cell(row=row_num, column=2).value or '').strip()
                manufacturer = str(ws.cell(row=row_num, column=3).value or '').strip()
                item_gtin = str(ws.cell(row=row_num, column=4).value or '').strip()
                tags_str = str(ws.cell(row=row_num, column=5).value or '').strip()
                items_per_box = ws.cell(row=row_num, column=6).value
                cost_per_item = ws.cell(row=row_num, column=7).value
                url = str(ws.cell(row=row_num, column=8).value or '').strip()
                notes_public = str(ws.cell(row=row_num, column=9).value or '').strip()
                notes_private = str(ws.cell(row=row_num, column=10).value or '').strip()
                stock_id = str(ws.cell(row=row_num, column=11).value or '').strip()
                organization_name = str(ws.cell(row=row_num, column=12).value or '').strip()
                quantity = ws.cell(row=row_num, column=13).value
                location = str(ws.cell(row=row_num, column=14).value or '').strip()
                detail = str(ws.cell(row=row_num, column=15).value or '').strip()
                stock_gtin = str(ws.cell(row=row_num, column=16).value or '').strip()
                date_received = ws.cell(row=row_num, column=17).value
                expiration_date = ws.cell(row=row_num, column=18).value
                lot_number = str(ws.cell(row=row_num, column=19).value or '').strip()
                stock_notes = str(ws.cell(row=row_num, column=20).value or '').strip()
                surplus_status = str(ws.cell(row=row_num, column=21).value or '').strip()
                
                # Skip empty rows
                if not item_name:
                    continue
                
                # Validate item ID if provided
                if item_id and item_id not in existing_items:
                    errors.append(f"Row {row_num}: Item ID '{item_id}' does not exist")
                    continue
                
                # Validate stock item ID if provided
                if stock_id and stock_id not in existing_stock_items:
                    errors.append(f"Row {row_num}: Stock Item ID '{stock_id}' does not exist")
                    continue
                
                # Validate organization
                if not organization_name:
                    errors.append(f"Row {row_num}: Organization is required")
                    continue
                
                organization = organizations_dict.get(organization_name.lower())
                if not organization:
                    errors.append(f"Row {row_num}: Organization '{organization_name}' not found")
                    continue
                
                # Validate tags if provided
                valid_tags = []
                if tags_str:
                    tag_names = [name.strip() for name in tags_str.split(',') if name.strip()]
                    for tag_name in tag_names:
                        tag = tags_dict.get(tag_name.lower())
                        if not tag:
                            errors.append(f"Row {row_num}: Tag '{tag_name}' does not exist")
                            continue
                        valid_tags.append(tag)
                
                # Validate quantity
                try:
                    if quantity is not None:
                        quantity = int(quantity)
                        if quantity < 0:
                            errors.append(f"Row {row_num}: Quantity cannot be negative")
                            continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Quantity must be a valid number")
                    continue
                
                # Store row data for session
                row_data = {
                    'item_id': item_id,
                    'item_name': item_name,
                    'manufacturer': manufacturer,
                    'item_gtin': item_gtin,
                    'tags_str': tags_str,
                    'items_per_box': items_per_box,
                    'cost_per_item': cost_per_item,
                    'url': url,
                    'notes_public': notes_public,
                    'notes_private': notes_private,
                    'stock_id': stock_id,
                    'organization_name': organization_name,
                    'quantity': quantity,
                    'location': location,
                    'detail': detail,
                    'stock_gtin': stock_gtin,
                    'date_received': date_received,
                    'expiration_date': expiration_date,
                    'lot_number': lot_number,
                    'stock_notes': stock_notes,
                    'surplus_status': surplus_status,
                    'valid_tags': [str(tag.id) for tag in valid_tags] if valid_tags else []
                }
                uploaded_data.append(row_data)
                
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing row - {str(e)}")
        
        # If there are validation errors, don't proceed with creation
        if errors:
            for error in errors[:10]:  # Show first 10 errors
                messages.error(request, error)
            if len(errors) > 10:
                messages.warning(request, f'... and {len(errors) - 10} more errors.')
            messages.error(request, f'Found {len(errors)} validation errors. Please fix them and try again.')
            return redirect('bulk_edit')
        
        # Store uploaded data in session for the preview page
        request.session['uploaded_inventory_data'] = json.dumps(uploaded_data, default=json_serial)
        
        # Redirect to the preview page
        return redirect('bulk_edit_preview')
        
    except Exception as e:
        messages.error(request, f'Error processing file: {str(e)}')
        return redirect('bulk_edit')


@login_required
@permission_required('inventory.bulkedit_item', raise_exception=True)
def preview_changes(request):
    """
    Preview changes from uploaded spreadsheet before executing them.
    
    Returns:
        HttpResponse: Rendered template with preview of changes
    """
    # Check if we have uploaded data in the session
    if 'uploaded_inventory_data' not in request.session:
        messages.error(request, 'No uploaded data found. Please upload a spreadsheet first.')
        return redirect('bulk_edit')
    
    try:
        # Load the uploaded data from the session and parse date fields
        uploaded_data = parse_date_fields(json.loads(request.session['uploaded_inventory_data']))
        
        # Prefetch all data to avoid N+1 queries
        existing_items = {str(item.id): item for item in Item.objects.all()}
        existing_stock_items = {str(stock.id): stock for stock in StockItem.objects.all()}
        
        # Track changes
        item_changes = []  # List of item changes
        stock_item_changes = []  # List of stock item changes
        stock_item_moves = []  # List of stock item moves
        new_items = []  # List of new items
        new_stock_items = []  # List of new stock items
        
        # Process each row in the uploaded data
        for row in uploaded_data:
            item_id = row.get('item_id')
            stock_id = row.get('stock_id')
            
            # Check if this is an existing item
            if item_id and item_id in existing_items:
                item = existing_items[item_id]
                
                # Check for item changes
                if row['item_name'] != item.name:
                    item_changes.append({
                        'item_name': item.name,
                        'field': 'Name',
                        'old_value': item.name,
                        'new_value': row['item_name']
                    })
                
                if row['manufacturer'] != item.manufacturer:
                    item_changes.append({
                        'item_name': item.name,
                        'field': 'Manufacturer',
                        'old_value': item.manufacturer,
                        'new_value': row['manufacturer']
                    })
                
                if row['item_gtin'] != item.gtin:
                    item_changes.append({
                        'item_name': item.name,
                        'field': 'GTIN',
                        'old_value': item.gtin,
                        'new_value': row['item_gtin']
                    })
                
                # Check for stock item changes
                if stock_id and stock_id in existing_stock_items:
                    stock = existing_stock_items[stock_id]
                    
                    # Check if stock item is being moved to another item
                    if str(stock.item.id) != item_id:
                        stock_item_moves.append({
                            'detail': stock.detail,
                            'quantity': stock.quantity,
                            'from_item': stock.item.name,
                            'to_item': row['item_name']
                        })
                    
                    # Check for stock item field changes
                    if int(row['quantity']) != stock.quantity:
                        stock_item_changes.append({
                            'item_name': item.name,
                            'detail': stock.detail,
                            'field': 'Quantity',
                            'old_value': stock.quantity,
                            'new_value': row['quantity']
                        })
                    
                    if row['location'] != stock.location:
                        stock_item_changes.append({
                            'item_name': item.name,
                            'detail': stock.detail,
                            'field': 'Location',
                            'old_value': stock.location,
                            'new_value': row['location']
                        })
                    
                    if row['detail'] != stock.detail:
                        stock_item_changes.append({
                            'item_name': item.name,
                            'detail': stock.detail,
                            'field': 'Detail',
                            'old_value': stock.detail,
                            'new_value': row['detail']
                        })
                else:
                    # This is a new stock item for an existing item
                    new_stock_items.append({
                        'item_name': row['item_name'],
                        'detail': row['detail'],
                        'quantity': row['quantity'],
                        'location': row['location']
                    })
            else:
                # This is a new item
                new_items.append({
                    'name': row['item_name'],
                    'manufacturer': row['manufacturer'],
                    'tags': row['tags_str'],
                    'gtin': row['item_gtin']
                })
        
        # Prepare context for the template
        context = {
            'changes_detected': bool(item_changes or stock_item_changes or stock_item_moves or new_items or new_stock_items),
            'item_changes': item_changes,
            'stock_item_changes': stock_item_changes,
            'stock_item_moves': stock_item_moves,
            'new_items': new_items,
            'new_stock_items': new_stock_items,
        }
        
        return render(request, 'bulk_edit/preview.html', context)
        
    except Exception as e:
        messages.error(request, f'Error processing preview: {str(e)}')
        return redirect('bulk_edit')


@login_required
@permission_required('inventory.bulkedit_item', raise_exception=True)
def execute_changes(request):
    """
    Execute changes from uploaded spreadsheet after preview.
    
    Returns:
        HttpResponse: Redirect to bulk edit page with success/error message
    """
    if request.method != 'POST':
        return redirect('bulk_edit')
    
    # Check if we have uploaded data in the session
    if 'uploaded_inventory_data' not in request.session:
        messages.error(request, 'No uploaded data found. Please upload a spreadsheet first.')
        return redirect('bulk_edit')
    
    try:
        # Load the uploaded data from the session and parse date fields
        uploaded_data = parse_date_fields(json.loads(request.session['uploaded_inventory_data']))
        
        # Prefetch all data to avoid N+1 queries
        existing_items = {str(item.id): item for item in Item.objects.all()}
        existing_stock_items = {str(stock.id): stock for stock in StockItem.objects.all()}
        organizations_dict = {org.name.lower(): org for org in Organization.objects.all()}
        
        # Track changes for audit logging
        changes_made = []
        
        # Execute all changes within a transaction
        with transaction.atomic():
            # Process each row in the uploaded data
            for row in uploaded_data:
                item_id = row.get('item_id')
                stock_id = row.get('stock_id')
                
                # Handle existing items
                if item_id and item_id in existing_items:
                    item = existing_items[item_id]
                    
                    # Check for item changes
                    item_changed = False
                    before_state = audit_log_state(item)
                    
                    if row['item_name'] != item.name:
                        item.name = row['item_name']
                        item_changed = True
                    
                    if row['manufacturer'] != item.manufacturer:
                        item.manufacturer = row['manufacturer']
                        item_changed = True
                    
                    if row['item_gtin'] != item.gtin:
                        item.gtin = row['item_gtin']
                        item_changed = True
                    
                    if row['url'] != item.url:
                        item.url = row['url']
                        item_changed = True
                    
                    if row['notes_public'] != item.notes_public:
                        item.notes_public = row['notes_public']
                        item_changed = True
                    
                    if row['notes_private'] != item.notes_private:
                        item.notes_private = row['notes_private']
                        item_changed = True
                    
                    # Save item if changed
                    if item_changed:
                        item.save()
                        after_state = audit_log_state(item)
                        audit_log_event(
                            request.user,
                            f"Bulk edited item '{item.name}'",
                            before_state,
                            after_state
                        )
                        changes_made.append(f"Updated item '{item.name}'")
                    
                    # Handle stock item changes
                    if stock_id and stock_id in existing_stock_items:
                        stock = existing_stock_items[stock_id]
                        stock_changed = False
                        before_state = audit_log_state(stock)
                        
                        # Check if stock item is being moved to another item
                        if str(stock.item.id) != item_id:
                            old_item_name = stock.item.name
                            stock.item = item
                            stock_changed = True
                            changes_made.append(f"Moved stock from '{old_item_name}' to '{item.name}'")
                        
                        # Check for other stock item changes
                        if int(row['quantity']) != stock.quantity:
                            stock.quantity = int(row['quantity'])
                            stock_changed = True
                        
                        if row['location'] != stock.location:
                            stock.location = row['location']
                            stock_changed = True
                        
                        if row['detail'] != stock.detail:
                            stock.detail = row['detail']
                            stock_changed = True
                        
                        if row['stock_gtin'] != stock.gtin:
                            stock.gtin = row['stock_gtin']
                            stock_changed = True
                        
                        if row['lot_number'] != stock.lot_number:
                            stock.lot_number = row['lot_number']
                            stock_changed = True
                        
                        if row['stock_notes'] != stock.notes:
                            stock.notes = row['stock_notes']
                            stock_changed = True
                        
                        if row['surplus_status'] != stock.surplus_status:
                            stock.surplus_status = row['surplus_status']
                            stock_changed = True
                        
                        # Save stock item if changed
                        if stock_changed:
                            stock.save()
                            after_state = audit_log_state(stock)
                            audit_log_event(
                                request.user,
                                f"Bulk edited stock item for '{item.name}'",
                                before_state,
                                after_state
                            )
                            changes_made.append(f"Updated stock item for '{item.name}'")
                    else:
                        # Create new stock item for existing item
                        organization = organizations_dict.get(row['organization_name'].lower())
                        if organization:
                            new_stock = StockItem(
                                item=item,
                                organization=organization,
                                quantity=int(row['quantity']),
                                location=row['location'],
                                detail=row['detail'],
                                gtin=row['stock_gtin'],
                                date_received=row['date_received'],
                                expiration_date=row['expiration_date'],
                                lot_number=row['lot_number'],
                                notes=row['stock_notes'],
                                surplus_status=row['surplus_status']
                            )
                            new_stock.save()
                            after_state = audit_log_state(new_stock)
                            audit_log_event(
                                request.user,
                                f"Created new stock item for '{item.name}' via bulk edit",
                                audit_log_state(None),
                                after_state
                            )
                            changes_made.append(f"Added new stock item to '{item.name}'")
                else:
                    # Create new item
                    organization = organizations_dict.get(row['organization_name'].lower())
                    if organization:
                        new_item = Item(
                            name=row['item_name'],
                            manufacturer=row['manufacturer'],
                            gtin=row['item_gtin'],
                            url=row['url'],
                            notes_public=row['notes_public'],
                            notes_private=row['notes_private']
                        )
                        new_item.save()
                        
                        # Add tags if provided
                        if row['valid_tags']:
                            tag_ids = row['valid_tags']
                            tags = Tag.objects.filter(id__in=tag_ids)
                            new_item.tags.set(tags)
                        
                        after_state = audit_log_state(new_item)
                        audit_log_event(
                            request.user,
                            f"Created new item '{new_item.name}' via bulk edit",
                            audit_log_state(None),
                            after_state
                        )
                        changes_made.append(f"Created new item '{new_item.name}'")
                        
                        # Create stock item for the new item
                        new_stock = StockItem(
                            item=new_item,
                            organization=organization,
                            quantity=int(row['quantity']),
                            location=row['location'],
                            detail=row['detail'],
                            gtin=row['stock_gtin'],
                            date_received=row['date_received'],
                            expiration_date=row['expiration_date'],
                            lot_number=row['lot_number'],
                            notes=row['stock_notes'],
                            surplus_status=row['surplus_status']
                        )
                        new_stock.save()
                        after_state = audit_log_state(new_stock)
                        audit_log_event(
                            request.user,
                            f"Created initial stock for '{new_item.name}' via bulk edit",
                            audit_log_state(None),
                            after_state
                        )
        
        # Show success message with number of changes made
        if changes_made:
            messages.success(request, f'Successfully applied {len(changes_made)} changes to the inventory.')
        else:
            messages.info(request, 'No changes were detected or applied to the inventory.')
        
        # Clear session data
        del request.session['uploaded_inventory_data']
        
        return redirect('bulk_edit')
        
    except Exception as e:
        messages.error(request, f'Error executing changes: {str(e)}')
        return redirect('bulk_edit_preview')
