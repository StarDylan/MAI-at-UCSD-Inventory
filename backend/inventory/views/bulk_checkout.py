"""
Bulk checkout views for managing multi-item checkouts with organization tracking.
"""

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from inventory.forms import CheckOutForm, CheckOutCompleteForm, CheckOutItemEditForm, CheckOutItemDetailEditForm
from inventory.models import Item, StockItem, AuditEvent, CheckOut, CheckOutItem, Location
from .utils import audit_log_state, audit_log_event


class BulkCheckoutListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Main bulk checkout page with Active and Completed tabs.
    """
    model = CheckOut
    template_name = 'checkout/bulk_checkout_list.html'
    context_object_name = 'checkouts'
    permission_required = 'inventory.view_checkout'
    paginate_by = 20
    
    def get_queryset(self):
        from django.db.models import Q

        tab = self.request.GET.get('tab', 'active')
        search_query = self.request.GET.get('search', '').strip()

        # Base queryset depending on tab
        if tab == 'completed':
            queryset = CheckOut.objects.filter(is_completed=True).order_by('-completed_at')
        else:
            queryset = CheckOut.objects.filter(is_completed=False).order_by('-created_at')

        # Add search filtering if provided
        if search_query:
            queryset = queryset.filter(
                Q(organization__name__icontains=search_query) |
                Q(notes__icontains=search_query) |
                Q(created_by__username__icontains=search_query) |
                Q(created_by__first_name__icontains=search_query) |
                Q(created_by__last_name__icontains=search_query)
            )

        # Always select related org/user and prefetch checkout_items and their stock/item
        queryset = queryset.select_related('organization', 'created_by').prefetch_related(
            'checkout_items__stock_item__item',
            'checkout_items__stock_item__organization',
        )

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_tab'] = self.request.GET.get('tab', 'active')
        context['search_query'] = self.request.GET.get('search', '')
        context['active_count'] = CheckOut.objects.filter(is_completed=False).count()
        context['completed_count'] = CheckOut.objects.filter(is_completed=True).count()
        return context


@login_required
@permission_required('inventory.add_checkout', raise_exception=True)
def checkout_create_view(request):
    """
    Create a new active checkout.
    """
    if request.method == 'POST':
        form = CheckOutForm(request.POST)
        if form.is_valid():
            checkout = form.save(commit=False)
            checkout.created_by = request.user
            checkout.save()
            
            # Log the creation
            donation_type = "donation" if checkout.is_donation else "non-donation"
            audit_log_event(
                request.user,
                f"Created new {donation_type} checkout for {checkout.organization.name}",
                audit_log_state(None),
                audit_log_state(checkout)
            )
            
            messages.success(request, f'New checkout created for {checkout.organization.name}')
            return redirect('checkout_detail', checkout_id=checkout.id)
    else:
        form = CheckOutForm()
    
    return render(request, 'checkout/checkout_create.html', {
        'form': form
    })


@login_required
@permission_required('inventory.view_checkout', raise_exception=True)
def checkout_detail_view(request, checkout_id):
    """
    View and edit individual checkout details.
    """
    checkout = get_object_or_404(
        CheckOut.objects
        .select_related('organization')
        .prefetch_related(
            'checkout_items__stock_item__item',
            'checkout_items__stock_item__organization'
        ),
        id=checkout_id
    )
    
    checkout_items = checkout.checkout_items.all()

    # Process items in Python to calculate boxes and remaining
    for item in checkout_items:
        if item.stock_item.item.items_per_box is not None:
            item.boxes = item.quantity // item.stock_item.item.items_per_box
            item.remaining = item.quantity % item.stock_item.item.items_per_box
        else:
            item.boxes = None
            item.remaining = None

    # Get audit events for the checkout with user details (using select_related on user)
    checkout_audit = AuditEvent.objects.filter(
        entity_id=checkout.id
    ).select_related('user').order_by('-created_at')

    # Prepare audit events for template display
    for event in checkout_audit:
        event.json_data = {
            'before': event.before,
            'after': event.after,
        }

    context = {
        'checkout': checkout,
        'checkout_items': checkout_items,
        'checkout_audit': checkout_audit,
        'can_edit': not checkout.is_completed,
    }

    return render(request, 'checkout/checkout_detail.html', context)

@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_add_item_view(request, checkout_id):
    """
    Add an item to an existing checkout.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    if checkout.is_completed:
        messages.error(request, 'Cannot add items to completed checkout')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        # Handle AJAX request for adding item
        stock_item_id = request.POST.get('stock_item_id')
        quantity = int(request.POST.get('quantity', 1))
        notes = request.POST.get('notes', '')
        
        try:
            stock_item = StockItem.objects.get(id=stock_item_id)
            
            # Check if already exists
            existing_item = CheckOutItem.objects.filter(
                checkout=checkout, 
                stock_item=stock_item
            ).first()
            
            if existing_item:
                return JsonResponse({
                    'success': False, 
                    'error': 'This stock item is already in the checkout'
                })
            
            # Check quantity availability
            if quantity > stock_item.quantity:
                return JsonResponse({
                    'success': False,
                    'error': f'Only {stock_item.quantity} units available'
                })
            
            # Create the checkout item
            checkout_item = CheckOutItem.objects.create(
                checkout=checkout,
                stock_item=stock_item,
                quantity=quantity,
                notes=notes
            )
            
            # Log the addition
            audit_log_event(
                request.user,
                f"Added {quantity} of \"{stock_item.item.name}\" from location \"{stock_item.location_new.name}\" to checkout for {checkout.organization.name}",
                audit_log_state(None),
                audit_log_state(checkout_item),
                str(checkout.id)
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Added {quantity}x {stock_item.item.name} to checkout'
            })
            
        except StockItem.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Stock item not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request - show form to add items
    items = Item.active_objects.all().order_by('name')
    return render(request, 'checkout/checkout_add_item.html', {
        'checkout': checkout,
        'items': items,
    })


@login_required
@permission_required('inventory.view_checkout', raise_exception=True)
def export_checkout_items_view(request, checkout_id):
    """
    Export checkout items to an Excel file.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    sheet_title = f"Checkout {checkout.id}"[:31]
    ws.title = sheet_title
    
    # Define the column headers
    headers = [
        'Requested Product',
        'Party',
        'QTY',
        'Status',
        'Date of Pickup',
        'Weight (lbs)',
        'Estimated Cost of Donation',
        'Donation from?'
    ]
    
    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Get all checkout items with related data
    checkout_items = CheckOutItem.objects.filter(checkout=checkout).select_related(
        'stock_item__item',
        'stock_item__organization'
    )
    
    # Add data rows
    for row_num, item in enumerate(checkout_items, 2):
        ws.cell(row=row_num, column=1, value=item.stock_item.item.name)  # Requested Product
        ws.cell(row=row_num, column=2, value=checkout.organization.name)  # Party
        ws.cell(row=row_num, column=3, value=item.quantity)  # QTY
        ws.cell(row=row_num, column=4, value='Completed' if checkout.is_completed else 'In Progress')  # Status
        ws.cell(row=row_num, column=5, value=checkout.completed_at.date() if checkout.completed_at else 'N/A')  # Date of Pickup
        # Weight (lbs) per item is not available; avoid repeating total checkout weight on every row
        ws.cell(row=row_num, column=6, value='N/A')  # Weight (lbs)
        
        # Estimated Cost of Donation (using item cost if available, otherwise empty)
        cost_per_item = item.stock_item.item.cost_per_item
        total_cost = float(cost_per_item * item.quantity) if cost_per_item else 'N/A'
        ws.cell(row=row_num, column=7, value=total_cost)
        
        # Show the original organization of the item in the Donation from? column
        ws.cell(row=row_num, column=8, value=item.stock_item.organization.name if item.stock_item.organization else 'N/A')  # Donation from?
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Create the HttpResponse object with the appropriate Excel header
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=checkout_{checkout.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_remove_item_view(request, checkout_id, item_id):
    """
    Remove an item from a checkout.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    checkout_item = get_object_or_404(CheckOutItem, id=item_id, checkout=checkout)
    
    if checkout.is_completed:
        messages.error(request, 'Cannot remove items from completed checkout')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        item_description = str(checkout_item)
        
        # Log the removal
        audit_log_event(
            request.user,
            f"Removed {checkout_item.quantity} of \"{checkout_item.stock_item.item.name}\" from location \"{checkout_item.stock_item.location_new.name}\" from checkout for {checkout.organization.name}",
            audit_log_state(checkout_item),
            audit_log_state(None),
            str(checkout.id)
        )
        
        checkout_item.delete()
        messages.success(request, f'Removed {item_description} from checkout')
    
    return redirect('checkout_detail', checkout_id=checkout.id)


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_edit_item_view(request, checkout_id, item_id):
    """
    Edit quantity and notes for a checkout item.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    checkout_item = get_object_or_404(CheckOutItem, id=item_id, checkout=checkout)
    
    if checkout.is_completed:
        messages.error(request, 'Cannot edit items in completed checkout')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        old_quantity = checkout_item.quantity
        form = CheckOutItemEditForm(request.POST, instance=checkout_item)
        if form.is_valid():
            new_quantity = form.cleaned_data['quantity']
            
            # Check if there's enough stock for the new quantity
            available_stock = checkout_item.stock_item.quantity
            if new_quantity > available_stock:
                messages.error(
                    request, 
                    f'Only {available_stock} units available for {checkout_item.stock_item.item.name}'
                )
                return redirect('checkout_detail', checkout_id=checkout.id)
            
            form.save()
            
            # Log the edit
            audit_log_event(
                request.user,
                f"Updated quantity for \"{checkout_item.stock_item.item.name}\" in location \"{checkout_item.stock_item.location_new.name}\" "
                f"in checkout for {checkout.organization.name} from {old_quantity} to {new_quantity}",
                audit_log_state(checkout_item),
                audit_log_state(checkout_item),
                entity_id=str(checkout.id)
            )
            
            messages.success(
                request, 
                f'Updated quantity for {checkout_item.stock_item.item.name} '
                f'from {old_quantity} to {new_quantity}'
            )
    
    return redirect('checkout_detail', checkout_id=checkout.id)


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_edit_item_detail_view(request, checkout_id, item_id):
    """
    Detailed edit page for checkout item with quantity and cost editing.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    checkout_item = get_object_or_404(CheckOutItem, id=item_id, checkout=checkout)
    
    if checkout.is_completed:
        messages.error(request, 'Cannot edit items in completed checkout')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        form = CheckOutItemDetailEditForm(request.POST, checkout_item=checkout_item)
        if form.is_valid():
            old_quantity = checkout_item.quantity
            old_cost = checkout_item.stock_item.item.cost_per_item
            
            new_quantity = form.cleaned_data['quantity']
            new_cost = form.cleaned_data['cost_per_item']
            new_notes = form.cleaned_data['notes']
            with transaction.atomic():
                # Update checkout item
                checkout_item.quantity = new_quantity
                checkout_item.notes = new_notes
                checkout_item.save()
                
                # Update item cost if provided and different
                if new_cost is not None and new_cost != old_cost:
                    before_item_state = audit_log_state(checkout_item.stock_item.item)
                    checkout_item.stock_item.item.cost_per_item = new_cost
                    checkout_item.stock_item.item.save()
                    
                    # Log cost change
                    audit_log_event(
                        request.user,
                        f"Updated value per item for \"{checkout_item.stock_item.item.name}\" "
                        f"from ${old_cost or 'None'} to ${new_cost}",
                        before_item_state,
                        audit_log_state(checkout_item.stock_item.item)
                    )
                
                # Log quantity change if different
                if new_quantity != old_quantity:
                    audit_log_event(
                        request.user,
                        f"Updated quantity for \"{checkout_item.stock_item.item.name}\" in location \"{checkout_item.stock_item.location_new.name}\" "
                        f"in checkout for {checkout.organization.name} from {old_quantity} to {new_quantity}",
                        audit_log_state(checkout_item),
                        audit_log_state(checkout_item),
                        entity_id=str(checkout.id)
                    )
            
            return redirect('checkout_detail', checkout_id=checkout.id)
    else:
        form = CheckOutItemDetailEditForm(checkout_item=checkout_item)
    
    # Calculate additional context
    available_stock = checkout_item.stock_item.quantity
    boxes_available = None
    is_approx = False
    if checkout_item.stock_item.item.items_per_box:
        # Round to 1 decimal
        is_approx = (available_stock % checkout_item.stock_item.item.items_per_box) != 0
        if is_approx:
            boxes_available = round((available_stock / checkout_item.stock_item.item.items_per_box), 1)
        else:
            boxes_available = int(available_stock / checkout_item.stock_item.item.items_per_box)


    context = {
        'checkout': checkout,
        'checkout_item': checkout_item,
        'form': form,
        'available_stock': available_stock,
        'boxes_available': boxes_available,
        "boxes_available_is_approx": is_approx,
    }

    return render(request, 'checkout/checkout_edit_item.html', context)


@login_required
@permission_required('inventory.complete_checkout', raise_exception=True)
def checkout_complete_view(request, checkout_id):
    """
    Complete a checkout - subtract stock quantities and mark as completed.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    if checkout.is_completed:
        messages.error(request, 'Checkout is already completed')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if not checkout.checkout_items.exists():  # type: ignore
        messages.error(request, 'Cannot complete checkout with no items')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        form = CheckOutCompleteForm(request.POST, checkout=checkout)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Process each checkout item with prefetched relations to avoid N+1 queries
                    for checkout_item in checkout.checkout_items.select_related('stock_item__item').all():  # type: ignore
                        stock_item = checkout_item.stock_item
                        
                        # Check if still enough quantity available
                        if stock_item.quantity < checkout_item.quantity:
                            messages.error(
                                request, 
                                f'Insufficient stock for {stock_item.item.name} from {stock_item.location_new.name}. '
                                f'Available: {stock_item.quantity}, Required: {checkout_item.quantity}'
                            )

                            raise Exception('Insufficient stock to complete checkout')
                        
                        # Log state before changes
                        before_state = audit_log_state(stock_item)
                        
                        # Subtract quantity
                        stock_item.quantity -= checkout_item.quantity
                        stock_item.save()
                        
                        # Log the stock change
                        audit_log_event(
                            request.user,
                            f"Checked-out {checkout_item.quantity} of \"{stock_item.item.name}\" from location \"{stock_item.location_new.name}\" to {checkout.organization.name}",
                            before_state,
                            audit_log_state(stock_item)
                        )
                    
                    # Mark checkout as completed
                    checkout.is_completed = True
                    checkout.completed_by = request.user
                    checkout.completed_at = timezone.now()
                    checkout.total_weight = form.cleaned_data.get('total_weight')
                    
                    completion_notes = form.cleaned_data.get('notes')
                    if completion_notes:
                        checkout.notes = f"{checkout.notes}\n\nCompletion notes: {completion_notes}".strip()
                    
                    checkout.save()
                    
                    # Log checkout completion
                    donation_type = "donation" if checkout.is_donation else "non-donation"
                    audit_log_event(
                        request.user,
                        f"Completed {donation_type} checkout for {checkout.organization.name} "
                        f"with {checkout.total_items_count} total items",
                        audit_log_state(None),
                        audit_log_state(checkout)
                    )
                    
                    return redirect('bulk_checkout_list')
            except Exception:
                # An error occurred, do not complete checkout
                return redirect('checkout_detail', checkout_id=checkout.id)
            
    else:
        form = CheckOutCompleteForm(checkout=checkout)
    
    # Check for problematic surplus items based on checkout type
    checkout_items = checkout.checkout_items.select_related('stock_item__item').all()
    
    # For donations: warn about wanted items (since they shouldn't be donated)
    # For non-donations: only warn about pending items (wanted items are fine for non-donations)
    if checkout.is_donation:
        problematic_items = [
            item for item in checkout_items 
            if item.stock_item.surplus_status in ['wanted', 'pending']
        ]
    else:
        # Non-donation: only warn about pending items
        problematic_items = [
            item for item in checkout_items 
            if item.stock_item.surplus_status == 'pending'
        ]
    
    has_surplus_items = len(problematic_items) > 0
    
    return render(request, 'checkout/checkout_complete.html', {
        'checkout': checkout,
        'form': form,
        'has_surplus_items': has_surplus_items,
        'surplus_items': problematic_items,
    })


@login_required
@permission_required('inventory.undo_checkout', raise_exception=True)
def checkout_undo_view(request, checkout_id):
    """
    Undo a completed checkout - restore stock quantities.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    if not checkout.is_completed:
        messages.error(request, 'Cannot undo checkout that is not completed')
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        with transaction.atomic():
            # Restore stock quantities with prefetched relations to avoid N+1 queries  
            for checkout_item in checkout.checkout_items.select_related('stock_item__item').all():  # type: ignore
                stock_item = checkout_item.stock_item
                
                # Log state before changes
                before_state = audit_log_state(stock_item)
                
                # Restore quantity
                stock_item.quantity += checkout_item.quantity
                stock_item.save()
                
                # Log the restoration
                audit_log_event(
                    request.user,
                    f"Returned {checkout_item.quantity} of \"{stock_item.item.name}\" to location \"{stock_item.location_new.name}\" (checkout undo from {checkout.organization.name})",
                    before_state,
                    audit_log_state(stock_item)
                )
            
            # Mark checkout as not completed (revert to active)
            checkout.is_completed = False
            checkout.completed_by = None
            checkout.completed_at = None
            checkout.save()
            
            # Log the undo
            donation_type = "donation" if checkout.is_donation else "non-donation"
            audit_log_event(
                request.user,
                f"Undid completed {donation_type} checkout for {checkout.organization.name}",
                audit_log_state(checkout),
                audit_log_state(None)
            )
            
            messages.success(
                request, 
                f'Checkout undone successfully! Stock quantities restored for '
                f'{checkout.total_items_count} items'
            )
    
    return redirect('checkout_detail', checkout_id=checkout.id)


@login_required
@permission_required('inventory.change_item', raise_exception=True)
def add_to_checkout_from_item_view(request, item_uuid):
    """
    Add stock items to an existing checkout from the item detail page.
    """
    item = get_object_or_404(Item, id=item_uuid)
    
    if request.method == 'POST':
        checkout_id = request.POST.get('checkout')
        stock_item_ids = request.POST.getlist('stock_item_ids')
        notes = request.POST.get('notes', '')

        if not checkout_id:
            messages.error(request, 'Please select a checkout.')
        elif not stock_item_ids:
            messages.error(request, 'Please select at least one stock item.')
        else:
            checkout = get_object_or_404(
                CheckOut.objects.select_related('organization'),
                id=checkout_id,
                is_completed=False
            )

            stock_items = StockItem.objects.filter(
                id__in=stock_item_ids,
                item=item,
                quantity__gt=0
            ).select_related('location_new')

            if stock_items.count() != len(stock_item_ids):
                messages.error(request, 'One or more selected stock items are invalid or unavailable.')
            else:
                added_count = 0
                updated_count = 0
                with transaction.atomic():
                    for stock_item in stock_items:
                        quantity_raw = request.POST.get(f'quantity_{stock_item.id}')
                        try:
                            quantity = int(quantity_raw)
                        except (TypeError, ValueError):
                            quantity = 0

                        if quantity <= 0:
                            continue
                        if quantity > stock_item.quantity:
                            quantity = stock_item.quantity

                        # Check if item already exists in checkout
                        checkout_item, created = CheckOutItem.objects.get_or_create(
                            checkout=checkout,
                            stock_item=stock_item,
                            defaults={'quantity': quantity, 'notes': notes}
                        )
                        
                        if created:
                            added_count += 1
                            audit_log_event(
                                request.user,
                                f"Added {quantity} of \"{item.name}\" from location \"{stock_item.location_new.name}\" to checkout for {checkout.organization.name}",
                                audit_log_state(None),
                                audit_log_state(checkout_item),
                                checkout.id
                            )
                        else:
                            # Update existing item
                            old_quantity = checkout_item.quantity
                            checkout_item.quantity = quantity
                            if notes:
                                checkout_item.notes = notes
                            checkout_item.save()
                            updated_count += 1
                            audit_log_event(
                                request.user,
                                f"Updated quantity of \"{item.name}\" from {old_quantity} to {quantity} in checkout for {checkout.organization.name}",
                                audit_log_state(checkout_item),
                                audit_log_state(checkout_item),
                                checkout.id
                            )

                if added_count > 0 or updated_count > 0:
                    msg_parts = []
                    if added_count > 0:
                        msg_parts.append(f'Added {added_count} stock item(s)')
                    if updated_count > 0:
                        msg_parts.append(f'Updated {updated_count} item(s)')
                    
                    messages.success(
                        request,
                        f'Added {added_count} stock item(s) for {item.name} to checkout for {checkout.organization.name}'
                    )
                    return redirect('checkout_detail', checkout_id=checkout.id)
                else:
                    messages.error(request, 'Please enter a quantity greater than 0 for at least one selected stock item.')
    
    # Get stock items for the template
    stock_items = item.stock_items.filter(quantity__gt=0).order_by(
        'detail', 'expiration_date', 'date_received'  # type: ignore
    )
    
    # Get active checkouts for the template
    checkouts = CheckOut.objects.filter(is_completed=False).select_related('organization').order_by('-created_at')
    
    return render(request, 'checkout/add_to_checkout_from_item.html', {
        'item': item,
        'stock_items': stock_items,
        'checkouts': checkouts,
    })


def get_stock_items_api(request, item_uuid):
    """
    API endpoint to get stock items for a specific item (for AJAX).
    Returns all stock items with quantity greater than 0, including surplus status information for flagging.
    """
    item = get_object_or_404(Item, id=item_uuid)
    
    # Include all stock items with quantity > 0, regardless of surplus status
    stock_items = item.stock_items.filter(
        quantity__gt=0
    ).order_by('detail', 'expiration_date', 'date_received')
    
    data = []
    for stock_item in stock_items:
        data.append({
            'id': str(stock_item.id),
            'detail': stock_item.detail,
            'location': stock_item.location_new.name,
            'quantity': stock_item.quantity,
            'organization': stock_item.organization.name,
            'expiration_date': stock_item.expiration_date.strftime('%Y-%m-%d') if stock_item.expiration_date else None,
            'surplus_status': stock_item.surplus_status,
            'surplus_status_display': stock_item.surplus_status_display,
            'display_name': f"{stock_item.detail} - {stock_item.location_new.name} ({stock_item.quantity} available)"
        })
    
    return JsonResponse({'stock_items': data})


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def get_active_checkouts_api(request):
    """
    API endpoint to get active checkouts for the current user or all active checkouts.
    Used by AJAX to populate checkout selection dropdowns.
    """
    checkouts = CheckOut.objects.filter(is_completed=False).select_related('organization').order_by('-created_at')
    
    data = []
    for checkout in checkouts:
        data.append({
            'id': str(checkout.id),
            'organization': checkout.organization.name,
            'created_at': checkout.created_at.strftime('%Y-%m-%d %H:%M'),
            'display_name': f"{checkout.organization.name} - {checkout.created_at.strftime('%Y-%m-%d %H:%M')}"
        })
    
    return JsonResponse({'checkouts': data})


@login_required
def get_checkout_items_api(request, checkout_id):
    """
    API endpoint to get existing items in a checkout.
    Returns a map of stock_item_id -> quantity for display in the form.
    """
    checkout = get_object_or_404(
        CheckOut.objects.select_related('organization'),
        id=checkout_id,
        is_completed=False
    )
    
    items = CheckOutItem.objects.filter(checkout=checkout).values('stock_item_id', 'quantity')
    
    data = {}
    for item in items:
        data[str(item['stock_item_id'])] = item['quantity']
    
    return JsonResponse({'existing_items': data})


@login_required
@permission_required('inventory.delete_checkout', raise_exception=True)
def checkout_delete_view(request, checkout_id):
    """
    Delete a checkout.
    Only allows deletion of active (non-completed) checkouts.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    # Prevent deletion of completed checkouts
    if checkout.is_completed:
        messages.error(request, "Cannot delete a completed checkout. You must undo it first.")
        return redirect('checkout_detail', checkout_id=checkout.id)
    
    if request.method == 'POST':
        organization_name = checkout.organization.name
        checkout_date = checkout.created_at.strftime('%Y-%m-%d %H:%M')
        
        # Delete the checkout (this will cascade to delete related CheckoutItems)
        checkout.delete()
        
        messages.success(request, f"Checkout for {organization_name} ({checkout_date}) has been deleted.")
        return redirect('bulk_checkout_list')
    
    # For GET requests, show confirmation page
    context = {
        'checkout': checkout,
        'items_count': checkout.total_items_count,
    }
    return render(request, 'checkout/checkout_delete_confirm.html', context)


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_bulk_add_item_view(request, checkout_id):
    """
    AJAX endpoint to bulk add all stock items from a specific Item to an existing checkout.
    Creates only 1 audit event for the item, not for each stock item.
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    if checkout.is_completed:
        return JsonResponse({
            'success': False,
            'error': 'Cannot add items to completed checkout'
        })
    
    if request.method == 'POST':
        # Handle AJAX request for bulk adding item
        item_id = request.POST.get('item_id')
        
        try:
            item = Item.objects.get(id=item_id)
            
            # Get all stock items for this item with quantity > 0
            total_available = item.stock_items.filter(quantity__gt=0).count()

            # Existing checkout items for this item
            existing_items = CheckOutItem.objects.filter(
                checkout=checkout,
                stock_item__item=item
            ).select_related('stock_item')

            # Exclude stock items already in this checkout
            stock_items_to_add = item.stock_items.filter(
                quantity__gt=0
            ).exclude(checkout_items__checkout=checkout)
            
            if total_available == 0:
                return JsonResponse({
                    'success': False,
                    'error': f'No stock items available for {item.name}'
                })
            
            if not stock_items_to_add.exists() and not existing_items.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'All available stock items from "{item.name}" are already in the checkout'
                })
            
            # Add all stock items to the checkout using bulk_create (single query)
            with transaction.atomic():
                updated_count = 0
                for checkout_item in existing_items:
                    max_qty = checkout_item.stock_item.quantity
                    if checkout_item.quantity < max_qty:
                        checkout_item.quantity = max_qty
                        checkout_item.save(update_fields=['quantity'])
                        updated_count += 1

                checkout_items = [
                    CheckOutItem(
                        checkout=checkout,
                        stock_item=stock_item,
                        quantity=stock_item.quantity,
                        notes=''
                    )
                    for stock_item in stock_items_to_add
                ]
                CheckOutItem.objects.bulk_create(checkout_items)
                added_count = len(checkout_items)
                
                # Log a single audit event for this item addition (not per stock item)
                audit_log_event(
                    request.user,
                    f"Bulk added all ({added_count} stock item(s)) from \"{item.name}\" to checkout for {checkout.organization.name}",
                    audit_log_state(None),
                    audit_log_state(item),
                    str(checkout.id)
                )
            
            message_parts = []
            if added_count > 0:
                message_parts.append(f'Added all {added_count} stock item(s) from {item.name} to checkout')
            if updated_count > 0:
                message_parts.append(f'Updated {updated_count} existing item(s) to max quantity')

            return JsonResponse({
                'success': True,
                'message': '. '.join(message_parts) if message_parts else 'Already added.'
            })
            
        except Item.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Item not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request - return error
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
@permission_required('inventory.change_checkout', raise_exception=True)
def checkout_bulk_add_location_view(request, checkout_id):
    """
    AJAX endpoint to bulk add all stock items from a specific Location to an existing checkout.
    Creates 1 audit event per Item (not per stock item).
    """
    checkout = get_object_or_404(CheckOut.objects.select_related('organization'), id=checkout_id)
    
    if checkout.is_completed:
        return JsonResponse({
            'success': False,
            'error': 'Cannot add items to completed checkout'
        })
    
    if request.method == 'POST':
        # Handle AJAX request for bulk adding location
        location_id = request.POST.get('location_id')
        
        try:
            location = Location.objects.get(id=location_id)
            
            # Get all stock items from this location with quantity > 0
            total_available = StockItem.objects.filter(
                location_new=location,
                quantity__gt=0
            ).count()

            existing_items = CheckOutItem.objects.filter(
                checkout=checkout,
                stock_item__location_new=location
            ).select_related('stock_item__item')

            # Exclude stock items already in this checkout
            stock_items_to_add = StockItem.objects.filter(
                location_new=location,
                quantity__gt=0
            ).exclude(checkout_items__checkout=checkout).select_related('item')
            
            if total_available == 0:
                return JsonResponse({
                    'success': False,
                    'error': f'No stock items available in location {location.name}'
                })
            
            if not stock_items_to_add.exists() and not existing_items.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'All available stock items from location "{location.name}" are already in the checkout'
                })
            
            # Add all stock items from this location to the checkout using bulk_create (single query)
            with transaction.atomic():
                from django.db.models import Count

                updated_count = 0
                for checkout_item in existing_items:
                    max_qty = checkout_item.stock_item.quantity
                    if checkout_item.quantity < max_qty:
                        checkout_item.quantity = max_qty
                        checkout_item.save(update_fields=['quantity'])
                        updated_count += 1
                
                checkout_items = [
                    CheckOutItem(
                        checkout=checkout,
                        stock_item=stock_item,
                        quantity=stock_item.quantity,
                        notes=''
                    )
                    for stock_item in stock_items_to_add
                ]
                CheckOutItem.objects.bulk_create(checkout_items)
                total_stock_items = len(checkout_items)
                
                # Get aggregated count per item using single query (no loop over stock items)
                item_counts = (
                    stock_items_to_add
                    .values('item__id', 'item__name')
                    .annotate(count=Count('id'))
                    .order_by()
                )
                
                # Create audit events from aggregated data
                # Build a map of item_id to item instance for quick lookup
                item_map = {str(si.item.id): si.item for si in stock_items_to_add}
                unique_items_count = 0
                
                for item_data in item_counts:
                    unique_items_count += 1
                    item_id = str(item_data['item__id'])
                    item_name = item_data['item__name']
                    count = item_data['count']
                    
                    # Get item from our map (already in memory)
                    item = item_map.get(item_id)
                    if item:
                        audit_log_event(
                            request.user,
                            f"Bulk added all ({count} stock item(s)) of \"{item_name}\" from location \"{location.name}\" to checkout for {checkout.organization.name}",
                            audit_log_state(None),
                            audit_log_state(item),
                            str(checkout.id)
                        )
            message_parts = []
            if total_stock_items > 0:
                message_parts.append(
                    f'Added all {total_stock_items} stock item(s) from location "{location.name}" to checkout'
                )
            if unique_items_count > 0:
                message_parts.append(f'{unique_items_count} unique items')
            if updated_count > 0:
                message_parts.append(f'Updated {updated_count} existing item(s) to max quantity')

            return JsonResponse({
                'success': True,
                'message': '. '.join(message_parts) if message_parts else 'Already added.'
            })
            
        except Location.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Location not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request - return error
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })