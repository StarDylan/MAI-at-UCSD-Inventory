"""
Surplus reporting views for generating and processing Excel reports.

This module handles surplus stock tracking reports including Excel export
and import functionality for bulk updates.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.db.models import Q
from django.db import transaction

from inventory.models import StockItem
from .utils import audit_log_state, audit_log_event


@login_required
@permission_required('inventory.view_stockitem', raise_exception=True)
def export_surplus_report(request):
    """
    Generate and download an Excel report of all stock items with surplus status.
    
    Returns:
        HttpResponse: Excel file download
    """
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Surplus Stock Report"
    
    # Headers
    headers = [
        'Stock Item ID', 'Item Name', 'Detail', 'Quantity', 'Location',
        'Organization', 'Date Received', 'Expiration Date', 'Lot Number',
        'Surplus Status', 'GTIN', 'Notes'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Get all stock items with positive quantity
    stock_items = StockItem.objects.filter(quantity__gt=0).select_related(
        'item', 'organization'
    ).order_by('item__name', 'detail', 'date_received')
    
    # Add data rows
    for row, stock in enumerate(stock_items, 2):
        data = [
            str(stock.id),
            stock.item.name,
            stock.detail or '',
            stock.quantity,
            stock.location,
            stock.organization.name,
            stock.date_received.strftime('%Y-%m-%d') if stock.date_received else '',
            stock.expiration_date.strftime('%Y-%m-%d') if stock.expiration_date else '',
            stock.lot_number or '',
            stock.surplus_status,
            stock.gtin or '',
            stock.notes or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # Color code surplus status
            if col == 10:  # Surplus Status column
                if stock.surplus_status == 'pending':
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                elif stock.surplus_status == 'wanted':
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                elif stock.surplus_status == 'not_wanted':
                    cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'surplus_stock_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


@login_required 
@permission_required('inventory.change_stockitem', raise_exception=True)
def upload_surplus_report(request):
    """
    Handle Excel file upload to update surplus status of stock items.
    
    Returns:
        HttpResponse: Rendered template or redirect
    """
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file was uploaded.')
            return render(request, 'surplus/upload_report.html')
        
        excel_file = request.FILES['excel_file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
            return render(request, 'surplus/upload_report.html')
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Validate headers
            expected_headers = [
                'Stock Item ID', 'Item Name', 'Detail', 'Quantity', 'Location',
                'Organization', 'Date Received', 'Expiration Date', 'Lot Number',
                'Surplus Status', 'GTIN', 'Notes'
            ]
            
            actual_headers = [cell.value for cell in ws[1]]
            if actual_headers != expected_headers:
                messages.error(request, 'Invalid file format. Please use the template from the export function.')
                return render(request, 'surplus/upload_report.html')
            
            # Process data rows within a transaction
            updated_count = 0
            error_count = 0
            errors = []
            
            # First pass: validate all data
            updates_to_apply = []
            
            for row_num in range(2, ws.max_row + 1):
                try:
                    stock_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                    surplus_status = str(ws.cell(row=row_num, column=10).value or '').strip().lower()
                    
                    # Skip empty rows
                    if not stock_id:
                        continue
                    
                    # Validate surplus status
                    if surplus_status not in ['pending', 'wanted', 'not_wanted']:
                        errors.append(f"Row {row_num}: Invalid surplus status '{surplus_status}'. Must be 'pending', 'wanted', or 'not_wanted'.")
                        error_count += 1
                        continue
                    
                    # Validate stock item exists
                    try:
                        stock_item = StockItem.objects.get(id=stock_id)
                        old_status = stock_item.surplus_status
                        
                        # Only add to updates if status is actually changing
                        if old_status != surplus_status:
                            updates_to_apply.append({
                                'stock_item': stock_item,
                                'old_status': old_status,
                                'new_status': surplus_status,
                                'row_num': row_num
                            })
                            
                    except StockItem.DoesNotExist:
                        errors.append(f"Row {row_num}: Stock item with ID '{stock_id}' not found.")
                        error_count += 1
                        continue
                        
                except Exception as e:
                    errors.append(f"Row {row_num}: Error processing row - {str(e)}")
                    error_count += 1
            
            # If there are validation errors, don't proceed with updates
            if error_count > 0:
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.warning(request, f'... and {len(errors) - 10} more errors.')
                messages.error(request, f'Found {error_count} validation errors. No updates were applied.')
                return redirect('upload_surplus_report')
            
            # Apply all updates within a transaction
            try:
                with transaction.atomic():
                    for update in updates_to_apply:
                        stock_item = update['stock_item']
                        old_status = update['old_status']
                        new_status = update['new_status']
                        
                        # Log the state before the change for audit
                        before_state = audit_log_state(stock_item)
                        
                        stock_item.surplus_status = new_status
                        stock_item.save()
                        
                        updated_count += 1
                        
                        # Log the audit event
                        after_state = audit_log_state(stock_item)
                        audit_log_event(
                            request.user,
                            f"Updated surplus status for \"{stock_item.item.name}\" from {old_status} to {new_status} via Excel upload",
                            before_state,
                            after_state
                        )
                
                # Success message
                if updated_count > 0:
                    messages.success(request, f'Successfully updated {updated_count} stock items.')
                else:
                    messages.info(request, 'No changes were made. All surplus statuses were already up to date.')
                    
            except Exception as e:
                messages.error(request, f'Transaction failed: {str(e)}. No changes were applied.')
                return redirect('upload_surplus_report')
                
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('upload_surplus_report')
    
    return render(request, 'surplus/upload_report.html')


@login_required
@permission_required('inventory.view_stockitem', raise_exception=True) 
def surplus_summary(request):
    """
    Display a summary of surplus status for all stock items.
    
    Returns:
        HttpResponse: Rendered template with surplus statistics
    """
    # Get surplus status counts
    pending_count = StockItem.objects.filter(surplus_status='pending', quantity__gt=0).count()
    wanted_count = StockItem.objects.filter(surplus_status='wanted', quantity__gt=0).count()
    not_wanted_count = StockItem.objects.filter(surplus_status='not_wanted', quantity__gt=0).count()
    total_count = pending_count + wanted_count + not_wanted_count
    
    # Get recent stock items needing surplus review
    recent_pending = StockItem.objects.filter(
        surplus_status='pending', 
        quantity__gt=0
    ).select_related('item', 'organization').order_by('-date_received')[:10]
    
    context = {
        'pending_count': pending_count,
        'wanted_count': wanted_count,
        'not_wanted_count': not_wanted_count,
        'total_count': total_count,
        'recent_pending': recent_pending,
    }
    
    return render(request, 'surplus/summary.html', context)